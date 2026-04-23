"""
AS24 scraperis (Playwright).
1. Prisijungia prie AS24 portalo
2. Dyzelinas: eina i "Kaina pagal degaline", nuima LT filtra, eksportuoja Excel
3. AdBlue: kaina pagal sali
"""
import json
import os
import re
import tempfile
from datetime import datetime
from playwright.async_api import async_playwright
import openpyxl

import config

DEBUG_DIR = "debug"

AS24_PRICES_URL = "https://extranet.as24.com/extranet/lt/network/prices-by-station/multi-energy"


async def debug_screenshot(page, name):
    os.makedirs(DEBUG_DIR, exist_ok=True)
    path = os.path.join(DEBUG_DIR, f"{name}.png")
    try:
        await page.screenshot(path=path, full_page=False)
        print(f"[AS24] Screenshot: {path}")
    except Exception as e:
        print(f"[AS24] Screenshot klaida: {e}")


async def close_cookies(page):
    """Uzdaro cookies/consent popup jei yra."""
    try:
        try:
            await page.wait_for_selector(
                '#didomi-notice, [class*="consent"], [id*="consent"], [class*="didomi"], '
                '[class*="sp_message"], iframe[title*="consent" i], iframe[title*="privacy" i]',
                timeout=5000,
            )
            print("[AS24] Cookie popup aptiktas")
        except Exception:
            print("[AS24] Cookie popup nepasirade per 5s, bandome toliau")

        for selector in [
            '#didomi-notice-agree-button',
            '.didomi-continue-without-agreeing',
            '[data-testid="notice-agree-button"]',
        ]:
            btn = page.locator(selector)
            if await btn.count() > 0:
                await btn.first.click(force=True)
                print(f"[AS24] Cookies uzdarytas: {selector}")
                await page.wait_for_timeout(1000)
                return

        for text in ["Agree and close", "Continue without agreeing", "Accept all", "Accept", "Sutinku"]:
            btn = page.locator(
                f'button:has-text("{text}"), a:has-text("{text}"), '
                f'[role="button"]:has-text("{text}"), span:has-text("{text}")'
            )
            if await btn.count() > 0:
                await btn.first.click(force=True)
                print(f"[AS24] Cookies uzdarytas: '{text}'")
                await page.wait_for_timeout(1000)
                return

        for frame in page.frames:
            if frame == page.main_frame:
                continue
            for text in ["Agree and close", "Accept all", "Accept", "Continue without agreeing"]:
                try:
                    btn = frame.locator(f'button:has-text("{text}"), a:has-text("{text}")')
                    if await btn.count() > 0:
                        await btn.first.click(force=True)
                        print(f"[AS24] Cookies uzdarytas (iframe): '{text}'")
                        await page.wait_for_timeout(1000)
                        return
                except Exception:
                    pass

        print("[AS24] Cookies popup nerastas arba jau uzdarytas")
    except Exception as e:
        print(f"[AS24] Cookies klaida: {e}")


async def login_as24(page):
    """Prisijungia prie AS24 portalo."""
    print(f"[AS24] Einame i {config.AS24_URL}")
    await page.goto(config.AS24_URL, wait_until="networkidle", timeout=60000)
    await debug_screenshot(page, "01_as24_login_page")

    await close_cookies(page)
    await debug_screenshot(page, "02_as24_after_cookies")

    all_inputs = page.locator("input:visible")
    input_count = await all_inputs.count()
    print(f"[AS24] Matomi input laukai: {input_count}")

    if input_count >= 3:
        await all_inputs.nth(0).click()
        await all_inputs.nth(0).type(config.AS24_CLIENT_ID, delay=50)
        await all_inputs.nth(1).click()
        await all_inputs.nth(1).type(config.AS24_EMAIL, delay=50)
        await all_inputs.nth(2).click()
        await all_inputs.nth(2).type(config.AS24_PASSWORD, delay=50)
    else:
        text_inputs = page.locator('input[type="text"]:visible, input[type="number"]:visible, input:not([type]):visible')
        if await text_inputs.count() > 0:
            await text_inputs.first.click()
            await text_inputs.first.type(config.AS24_CLIENT_ID, delay=50)
        email_inputs = page.locator('input[type="email"]:visible, input[type="text"]:visible')
        if await email_inputs.count() > 1:
            await email_inputs.nth(1).click()
            await email_inputs.nth(1).type(config.AS24_EMAIL, delay=50)
        pwd_inputs = page.locator('input[type="password"]:visible')
        if await pwd_inputs.count() > 0:
            await pwd_inputs.first.click()
            await pwd_inputs.first.type(config.AS24_PASSWORD, delay=50)

    await debug_screenshot(page, "03_as24_credentials_filled")
    await page.wait_for_timeout(2000)  # reCAPTCHA laiko registruoti veiksmus

    submit = page.locator('button[type="submit"], button:has-text("PRISIJUNGIMAS"), button:has-text("Prisijungimas"), button:has-text("Login")')
    if await submit.count() > 0:
        await submit.first.click()
    else:
        buttons = page.locator("button:visible")
        btn_count = await buttons.count()
        if btn_count > 0:
            await buttons.last.click()

    await page.wait_for_load_state("networkidle", timeout=30000)
    await debug_screenshot(page, "04_as24_after_login")
    print(f"[AS24] URL po prisijungimo: {page.url}")


async def clear_country_filter(page):
    """Nuima salies filtra jei pasirinkta Lietuva (rodo visas salis)."""
    try:
        # Tikriname ar yra aktyvus filtras
        await page.wait_for_timeout(2000)

        # Bandome rasti "Lietuva" filtro zenkliuka ir spausti X
        # 1. Filtras gali buti kaip chip/tag su X mygtuku
        lietuva_chip = page.locator(
            'button:has-text("Lietuva"), [class*="chip"]:has-text("Lietuva"), '
            '[class*="tag"]:has-text("Lietuva"), [class*="badge"]:has-text("Lietuva")'
        )
        if await lietuva_chip.count() > 0:
            # Spaudziame ant paties chip (gali atidaryti dropdown)
            await lietuva_chip.first.click()
            await page.wait_for_timeout(1000)

        # 2. Dropdown viduje ieskom X prie Lietuva
        remove_btn = page.locator(
            '[class*="selected-item"] button, [class*="chip"] button[class*="remove"], '
            '[class*="chip"] button[class*="close"], [class*="tag"] button, '
            'button[aria-label*="remove" i], button[aria-label*="Lietuva" i]'
        )
        if await remove_btn.count() > 0:
            await remove_btn.first.click()
            print("[AS24] Lietuva filtras nuimtas")
            await page.wait_for_timeout(500)

        # 3. Taikome filtra
        taikyti = page.locator('button:has-text("TAIKYTI"), button:has-text("Taikyti")')
        if await taikyti.count() > 0:
            await taikyti.first.click()
            await page.wait_for_load_state("networkidle", timeout=15000)
            await page.wait_for_timeout(2000)
            print("[AS24] Filtras pritaikytas")

    except Exception as e:
        print(f"[AS24] Filtro nuemimo klaida (nekritine): {e}")


def build_station_map():
    """
    Sukuria stoteles pavadinimo => config atitikmenys zodynas.
    Naudoja export_name jei nurodytas, kitu atveju filter.
    """
    station_map = {}
    for s in config.AS24_DIESEL_STATIONS:
        export_name = s.get("export_name", s["filter"]).upper()
        station_map[export_name] = s
    return station_map


def parse_price_value(raw):
    """Konvertuoja kaina is teksto i float."""
    if raw is None:
        return None
    cleaned = re.sub(r"[^\d,.]", "", str(raw)).replace(",", ".")
    try:
        val = float(cleaned)
        return val if 0.3 < val < 5.0 else None
    except ValueError:
        return None


def parse_diesel_excel(file_path, today):
    """Parsina AS24 eksporto Excel faila ir grazina dyzelino kainu sarasa."""
    wb = openpyxl.load_workbook(file_path, data_only=True)
    ws = wb.active

    # Randame stulpeliu indeksus
    headers = [str(cell.value or "").strip() for cell in ws[1]]
    print(f"[AS24] Excel stulpeliai ({len(headers)}): {headers[:10]}")

    def find_col(keywords):
        for i, h in enumerate(headers):
            h_low = h.lower()
            if all(kw.lower() in h_low for kw in keywords):
                return i
        return None

    station_col = find_col(["degalin"])
    product_col = find_col(["produkt"])
    price_col   = find_col(["suma", "nuolaid"])

    print(f"[AS24] Stulpeliai: stotele={station_col}, produktas={product_col}, kaina={price_col}")
    if None in (station_col, product_col, price_col):
        raise ValueError(f"Nerasti butini stulpeliai Excel faile. Headers: {headers}")

    station_map = build_station_map()
    found = {}  # name -> price (kad neduobliuotum)

    for row in ws.iter_rows(min_row=2, values_only=True):
        station_raw = str(row[station_col] or "").strip().upper()
        product_raw = str(row[product_col] or "").strip().lower()

        if "gazole" not in product_raw and "diesel" not in product_raw:
            continue

        # Ieskome atitinkamos stoteles
        matched = None
        for export_name, cfg in station_map.items():
            if station_raw == export_name or export_name in station_raw:
                matched = cfg
                break

        if matched is None:
            continue

        name = matched["name"]
        if name in found:
            continue  # pirma rastas eilutes imame

        price = parse_price_value(row[price_col])
        if price:
            found[name] = price
            print(f"[AS24] {name}: {price} EUR/l (stotele: {station_raw})")

    results = []
    for s in config.AS24_DIESEL_STATIONS:
        price = found.get(s["name"])
        if price is None:
            print(f"[AS24] {s['name']}: nerasta")
        results.append({
            "name": s["name"],
            "country": s["country"],
            "date": today,
            "price": price,
        })

    return results


async def scrape_diesel_via_export(page, today):
    """Atsisiuncia AS24 kainu Excel eksporta ir istraukia dyzelino kainas."""
    print(f"[AS24] Einame i kainu puslapi: {AS24_PRICES_URL}")
    await page.goto(AS24_PRICES_URL, wait_until="networkidle", timeout=30000)
    await page.wait_for_timeout(2000)
    await debug_screenshot(page, "05_as24_prices_page")

    # Nuimame LT filtra jei yra
    await clear_country_filter(page)
    await debug_screenshot(page, "06_as24_filter_cleared")

    # Eksportuojame
    export_btn = page.locator('button:has-text("Eksportuoti"), a:has-text("Eksportuoti")')
    export_count = await export_btn.count()
    print(f"[AS24] 'Eksportuoti' mygtuku: {export_count}")

    if export_count == 0:
        await debug_screenshot(page, "07_as24_no_export_btn")
        raise ValueError("Nerasta 'Eksportuoti' mygtuko")

    with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as f:
        tmp_path = f.name

    try:
        async with page.expect_download(timeout=30000) as dl_info:
            await export_btn.first.click()
        download = await dl_info.value
        await download.save_as(tmp_path)
        print(f"[AS24] Failas issaugotas laikinai: {tmp_path}")
        await debug_screenshot(page, "07_as24_after_export")

        results = parse_diesel_excel(tmp_path, today)
    finally:
        if os.path.exists(tmp_path):
            os.unlink(tmp_path)

    return results


async def scrape_adblue_country(page, country):
    try:
        print(f"[AS24] Ieskom AdBlue: {country['name']}")

        by_country = page.locator('a:has-text("pagal sal"), button:has-text("pagal sal"), a:has-text("by country")')
        if await by_country.count() > 0:
            await by_country.first.click()
            await page.wait_for_load_state("networkidle", timeout=10000)

        await page.wait_for_timeout(2000)

        country_select = page.locator('select:near(:text("Salis")), select:near(:text("Country"))')
        if await country_select.count() > 0:
            await country_select.first.select_option(label=country["country_filter"])

        product_select = page.locator('select:near(:text("Degalai")), select:near(:text("Product"))')
        if await product_select.count() > 0:
            await product_select.first.select_option(label="AdBlue")

        await page.wait_for_timeout(2000)

        price = await extract_adblue_price(page, country["zone"])
        print(f"[AS24] {country['name']}: AdBlue = {price}")
        return price
    except Exception as e:
        print(f"[AS24] Klaida su {country['name']}: {e}")
        return None


async def extract_adblue_price(page, target_zone):
    rows = await page.query_selector_all("tr")
    for row in rows:
        text = await row.inner_text()
        if "adblue" in text.lower():
            if target_zone and target_zone not in text:
                continue
            cells = await row.query_selector_all("td")
            for cell in cells:
                cell_text = (await cell.inner_text()).strip()
                val = parse_price_value(cell_text)
                if val and 0.1 < val < 3.0:
                    return val
    return None


async def scrape_as24():
    today = datetime.now().strftime("%Y-%m-%d")
    diesel_results = []
    adblue_results = []

    async with async_playwright() as p:
        browser = await p.chromium.launch(headless=True)
        context = await browser.new_context(
            viewport={"width": 1280, "height": 800},
            user_agent="Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 Chrome/131.0.0.0 Safari/537.36",
            accept_downloads=True,
        )
        page = await context.new_page()

        try:
            await login_as24(page)

            # Dyzelino kainos per eksporta
            diesel_results = await scrape_diesel_via_export(page, today)

            # AdBlue kainos (tolimesnis zingsnis)
            for country in config.AS24_ADBLUE_COUNTRIES:
                price = await scrape_adblue_country(page, country)
                adblue_results.append({
                    "name": country["name"],
                    "code": country["code"],
                    "date": today,
                    "price": price,
                })

        except Exception as e:
            print(f"[AS24] Bendra klaida: {e}")
            await debug_screenshot(page, "99_as24_error")
            raise
        finally:
            await browser.close()

    return {"diesel": diesel_results, "adblue": adblue_results}


def run_as24_scraper():
    import asyncio
    return asyncio.run(scrape_as24())


if __name__ == "__main__":
    print("=== AS24 Scraper testas ===")
    try:
        result = run_as24_scraper()
        print(json.dumps(result, indent=2, ensure_ascii=False))
    except Exception as e:
        print(f"Klaida: {e}")
